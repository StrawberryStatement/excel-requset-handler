from __future__ import annotations

import json
import shutil
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from backend.app.commenting.schemas import CommentContext, DEFAULT_COMMENT_TEMPLATE
from backend.app.commenting.template_engine import render_comment
from backend.app.models import DEFAULT_SERVICE_COLUMN, CellValue, ExcelTable, WorkbookConfig
from backend.app.services.excel_io import load_excel_table, normalize_cell


SOURCE_LIST_TYPE = "\u9700\u6c42\u5217\u8868"
BASELINE_SOURCE = "\u57fa\u51c6\u9700\u6c42\u5217\u8868"


@dataclass(frozen=True)
class V2WorkflowConfig:
    workbook: WorkbookConfig = field(default_factory=WorkbookConfig)
    owned_services: tuple[str, ...] = ()
    comment_fields: tuple[str, ...] = ()
    comment_template: str = DEFAULT_COMMENT_TEMPLATE
    comment_target: str = "both"
    sync_linked: bool = False
    fill_defaults: dict[str, str] = field(default_factory=dict)
    source_type: str = SOURCE_LIST_TYPE


@dataclass
class NormalizedSourceRow:
    source_name: str
    source_type: str
    row_number: int
    raw: dict[str, CellValue]
    normalized: dict[str, CellValue]
    extension_columns: list[str]
    fe_id: str = ""
    rr_id: str = ""
    service: str = ""
    title: str = ""
    candidate_fe: str = ""
    relation_tags: set[str] = field(default_factory=set)
    issue_tags: set[str] = field(default_factory=set)
    severity: str = "\u65e0"
    fill_notes: list[str] = field(default_factory=list)


@dataclass
class V2WorkflowResult:
    job_id: str
    output_dir: Path
    package_path: Path
    sync_plan_json_path: Path
    response_payload: dict[str, Any]


def run_v2_workflow(
    baseline_path: Path,
    source_paths: list[Path],
    output_root: Path,
    config: V2WorkflowConfig | None = None,
) -> V2WorkflowResult:
    config = config or V2WorkflowConfig()
    job_id = uuid.uuid4().hex[:12]
    job_dir = output_root / job_id
    input_dir = job_dir / "input"
    output_dir = job_dir / "output"
    input_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)

    saved_baseline = input_dir / baseline_path.name
    shutil.copy2(baseline_path, saved_baseline)
    saved_sources: list[Path] = []
    for path in source_paths:
        target = input_dir / path.name
        shutil.copy2(path, target)
        saved_sources.append(target)

    baseline = load_excel_table(saved_baseline)
    _validate_baseline(baseline, config.workbook)
    source_tables = [load_excel_table(path) for path in saved_sources]

    rows = _normalize_source_rows(baseline, source_tables, config)
    _apply_relationship_rules(rows, baseline, config)

    preview_rows = _build_preview_rows(rows, baseline, config)
    rr_rows = _build_rr_rows(rows)
    issues = _build_issue_rows(rows)
    extension_rows = _build_extension_rows(rows)
    pending_fe_rows = _build_pending_fe_rows(rows)
    sync_comments = _build_sync_comments(rows, config)
    summary = _build_summary(preview_rows, rr_rows, rows, sync_comments)

    package_path = output_dir / "\u9700\u6c42\u6c47\u603b\u5305.xlsx"
    _write_package(
        package_path=package_path,
        summary=summary,
        preview_rows=preview_rows,
        rr_rows=rr_rows,
        extension_rows=extension_rows,
        sync_comments=sync_comments,
        pending_fe_rows=pending_fe_rows,
        issue_rows=issues,
        raw_rows=_build_raw_rows(rows, config),
    )

    sync_plan_json_path = output_dir / "sync_plan.json"
    sync_plan_json_path.write_text(json.dumps(sync_comments, ensure_ascii=False, indent=2), encoding="utf-8")

    response_payload = {
        "job_id": job_id,
        "summary": summary,
        "preview_rows": preview_rows,
        "rr_rows": rr_rows,
        "issues": issues,
        "filters": _build_filters(preview_rows),
        "downloads": {
            "package": f"/api/jobs/{job_id}/downloads/\u9700\u6c42\u6c47\u603b\u5305.xlsx",
            "sync_plan_json": f"/api/jobs/{job_id}/downloads/sync_plan.json",
            "master": f"/api/jobs/{job_id}/downloads/\u9700\u6c42\u6c47\u603b\u5305.xlsx",
        },
    }
    (output_dir / "job_summary.json").write_text(json.dumps(response_payload, ensure_ascii=False, indent=2), encoding="utf-8")

    return V2WorkflowResult(
        job_id=job_id,
        output_dir=output_dir,
        package_path=package_path,
        sync_plan_json_path=sync_plan_json_path,
        response_payload=response_payload,
    )


def _validate_baseline(baseline: ExcelTable, config: WorkbookConfig) -> None:
    required = [config.fe_column, config.service_column, config.title_column]
    missing = [column for column in required if column not in baseline.headers]
    if missing:
        raise ValueError(f"baseline missing required columns: {', '.join(missing)}")

    seen: set[str] = set()
    duplicates: set[str] = set()
    for row in baseline.records:
        fe_id = normalize_cell(row.get(config.fe_column))
        if not fe_id:
            continue
        if fe_id in seen:
            duplicates.add(fe_id)
        seen.add(fe_id)
    if duplicates:
        raise ValueError(f"Duplicate baseline FE values: {', '.join(sorted(duplicates))}")


def _normalize_source_rows(
    baseline: ExcelTable,
    source_tables: list[ExcelTable],
    config: V2WorkflowConfig,
) -> list[NormalizedSourceRow]:
    rows: list[NormalizedSourceRow] = []
    baseline_headers = set(baseline.headers)
    cfg = config.workbook

    for table in source_tables:
        extension_columns = [header for header in table.headers if header not in baseline_headers]
        service_column_missing = cfg.service_column not in table.headers
        for index, raw_row in enumerate(table.records, start=2):
            normalized = dict(raw_row)
            fill_notes: list[str] = []
            for field, default_value in config.fill_defaults.items():
                had_column = field in normalized
                if not had_column or not normalize_cell(normalized.get(field)):
                    normalized[field] = default_value
                    note = f"{field}\u7531\u7a7a\u503c\u586b\u5145" if had_column else f"{field}\u7531\u865a\u62df\u5217\u586b\u5145"
                    fill_notes.append(note)

            row = NormalizedSourceRow(
                source_name=table.source_name,
                source_type=config.source_type or SOURCE_LIST_TYPE,
                row_number=index,
                raw=dict(raw_row),
                normalized=normalized,
                extension_columns=extension_columns,
                fe_id=normalize_cell(normalized.get(cfg.fe_column)),
                rr_id=normalize_cell(normalized.get(cfg.rr_column)),
                service=normalize_cell(normalized.get(cfg.service_column)),
                title=normalize_cell(normalized.get(cfg.title_column)),
                fill_notes=fill_notes,
            )

            if service_column_missing and cfg.service_column not in config.fill_defaults:
                _add_issue(row, "\u7f3a\u4e91\u670d\u52a1\u5217", "\u4e25\u91cd")
            elif service_column_missing and cfg.service_column in config.fill_defaults:
                _add_issue(row, "\u7f3a\u4e91\u670d\u52a1\u5217\u5df2\u586b\u5145", "\u63d0\u9192")
            if not row.service:
                _add_issue(row, "\u4e91\u670d\u52a1\u4e3a\u7a7a", "\u4e25\u91cd")
            if not row.fe_id and not row.rr_id:
                _add_issue(row, "\u7f3a\u4e3b\u952e", "\u4e25\u91cd")

            rows.append(row)
    return rows


def _apply_relationship_rules(rows: list[NormalizedSourceRow], baseline: ExcelTable, config: V2WorkflowConfig) -> None:
    cfg = config.workbook
    baseline_by_fe = baseline.by_key(cfg.fe_column)
    rr_to_fes: dict[str, set[str]] = {}
    fe_to_rrs: dict[str, set[str]] = {}
    fe_to_services: dict[str, set[str]] = {}
    fe_to_sources: dict[str, set[str]] = {}

    for row in rows:
        if row.rr_id and row.fe_id:
            rr_to_fes.setdefault(row.rr_id, set()).add(row.fe_id)
            fe_to_rrs.setdefault(row.fe_id, set()).add(row.rr_id)
        if row.fe_id and row.service:
            fe_to_services.setdefault(row.fe_id, set()).add(row.service)
            fe_to_sources.setdefault(row.fe_id, set()).add(row.source_name)

    for base_row in baseline.records:
        fe_id = normalize_cell(base_row.get(cfg.fe_column))
        rr_id = normalize_cell(base_row.get(cfg.rr_column))
        service = normalize_cell(base_row.get(cfg.service_column))
        if rr_id and fe_id:
            rr_to_fes.setdefault(rr_id, set()).add(fe_id)
            fe_to_rrs.setdefault(fe_id, set()).add(rr_id)
        if fe_id and service:
            fe_to_services.setdefault(fe_id, set()).add(service)

    for row in rows:
        if row.rr_id and not row.fe_id:
            candidates = sorted(rr_to_fes.get(row.rr_id, set()))
            if len(candidates) == 1:
                row.candidate_fe = candidates[0]
                row.relation_tags.add("FE\u5019\u9009")
                _add_issue(row, "FE\u5019\u9009", "\u63d0\u9192")
            else:
                row.relation_tags.add("\u5f85\u5efaFE")
                _add_issue(row, "\u5f85\u5efaFE", "\u63d0\u9192")

        if row.fe_id and not row.rr_id:
            row.relation_tags.add("\u65e0\u5916\u90e8RR")

        if row.fe_id and row.fe_id not in baseline_by_fe:
            _add_issue(row, "\u57fa\u51c6\u5916FE", "\u63d0\u9192")

        if row.rr_id and len(rr_to_fes.get(row.rr_id, set())) > 1:
            row.relation_tags.add("RR\u62c6\u5206")

        if row.fe_id and len(fe_to_rrs.get(row.fe_id, set())) > 1:
            row.relation_tags.add("RR\u5408\u5e76")

        if row.fe_id and len(fe_to_services.get(row.fe_id, set())) > 1:
            row.relation_tags.add("\u4e91\u670d\u52a1\u51b2\u7a81")
            _add_issue(row, "\u4e91\u670d\u52a1\u51b2\u7a81", "\u4e25\u91cd")

        if row.fe_id and len(fe_to_sources.get(row.fe_id, set())) > 1:
            row.relation_tags.add("\u591a\u6765\u6e90\u5171\u540c\u7ba1\u7406")
            _add_issue(row, "\u591a\u6765\u6e90\u5171\u540c\u7ba1\u7406", "\u9ad8\u4f18\u5148\u7ea7\u5f02\u5e38")

        _detect_baseline_changes(row, baseline_by_fe, cfg)


def _detect_baseline_changes(row: NormalizedSourceRow, baseline_by_fe: dict[str, dict[str, CellValue]], cfg: WorkbookConfig) -> None:
    if not row.fe_id or row.fe_id not in baseline_by_fe:
        return
    baseline_row = baseline_by_fe[row.fe_id]
    ignored = {cfg.fe_column, cfg.rr_column, cfg.service_column, *cfg.ignored_compare_columns}
    for field, baseline_value in baseline_row.items():
        if field in ignored or field not in row.normalized:
            continue
        source_value = row.normalized.get(field)
        if normalize_cell(baseline_value) != normalize_cell(source_value):
            _add_issue(row, "\u57fa\u51c6\u5b57\u6bb5\u53d8\u66f4", "\u63d0\u9192")
            return


def _add_issue(row: NormalizedSourceRow, issue: str, severity: str) -> None:
    row.issue_tags.add(issue)
    row.severity = _max_severity(row.severity, severity)


def _max_severity(left: str, right: str) -> str:
    order = {"\u65e0": 0, "\u4fe1\u606f": 1, "\u63d0\u9192": 2, "\u9ad8\u4f18\u5148\u7ea7\u5f02\u5e38": 3, "\u4e25\u91cd": 4}
    return left if order.get(left, 0) >= order.get(right, 0) else right


def _build_preview_rows(rows: list[NormalizedSourceRow], baseline: ExcelTable, config: V2WorkflowConfig) -> list[dict[str, Any]]:
    cfg = config.workbook
    grouped: dict[str, dict[str, Any]] = {}

    for row in rows:
        key = row.fe_id or row.candidate_fe or f"RR:{row.rr_id}" or f"{row.source_name}:{row.row_number}"
        item = grouped.setdefault(key, {
            "key": key,
            "fe_ids": set(),
            "rr_ids": set(),
            "services": set(),
            "sources": set(),
            "titles": set(),
            "relation_tags": set(),
            "issue_tags": set(),
            "severity": "\u65e0",
            "extension_summary": [],
            "owned": False,
        })
        if row.fe_id or row.candidate_fe:
            item["fe_ids"].add(row.fe_id or row.candidate_fe)
        if row.rr_id:
            item["rr_ids"].add(row.rr_id)
        if row.service:
            item["services"].add(row.service)
        if row.source_name:
            item["sources"].add(row.source_name)
        if row.title:
            item["titles"].add(row.title)
        item["relation_tags"].update(row.relation_tags)
        item["issue_tags"].update(row.issue_tags)
        item["severity"] = _max_severity(item["severity"], row.severity)
        item["owned"] = item["owned"] or row.service in config.owned_services
        summary = _extension_summary(row, config)
        if summary:
            item["extension_summary"].append(f"[{row.source_name}] {summary}")

    for base_row in baseline.records:
        fe_id = normalize_cell(base_row.get(cfg.fe_column))
        if not fe_id or fe_id in grouped:
            continue
        service = normalize_cell(base_row.get(cfg.service_column))
        rr_id = normalize_cell(base_row.get(cfg.rr_column))
        title = normalize_cell(base_row.get(cfg.title_column))
        grouped[fe_id] = {
            "key": fe_id,
            "fe_ids": {fe_id},
            "rr_ids": {rr_id} if rr_id else set(),
            "services": {service} if service else set(),
            "sources": set(),
            "titles": {title} if title else set(),
            "relation_tags": set(),
            "issue_tags": set(),
            "severity": "\u65e0",
            "extension_summary": [],
            "owned": service in config.owned_services,
        }

    output: list[dict[str, Any]] = []
    for item in grouped.values():
        output.append({
            "key": item["key"],
            "fe_id": ", ".join(sorted(item["fe_ids"])),
            "rr_id": ", ".join(sorted(item["rr_ids"])),
            "service": ", ".join(sorted(item["services"])),
            "source": ", ".join(sorted(item["sources"])),
            "title": " / ".join(sorted(item["titles"])),
            "owned": "\u662f" if item["owned"] else "\u5426",
            "relation": ", ".join(sorted(item["relation_tags"])) or "\u6b63\u5e38",
            "issue": ", ".join(sorted(item["issue_tags"])) or "\u65e0",
            "severity": item["severity"],
            "extension_summary": "\n".join(item["extension_summary"]),
        })
    return sorted(output, key=lambda item: (item["severity"], item["service"], item["fe_id"], item["rr_id"]))


def _extension_summary(row: NormalizedSourceRow, config: V2WorkflowConfig) -> str:
    fields = list(config.comment_fields) or row.extension_columns
    pieces = []
    for field in fields:
        value = normalize_cell(row.normalized.get(field))
        if value:
            pieces.append(f"{field}: {value}")
    return "; ".join(pieces)


def _build_rr_rows(rows: list[NormalizedSourceRow]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, set[str]]] = {}
    for row in rows:
        if not row.rr_id:
            continue
        item = grouped.setdefault(row.rr_id, {
            "fes": set(),
            "services": set(),
            "sources": set(),
            "relations": set(),
            "issues": set(),
        })
        if row.fe_id or row.candidate_fe:
            item["fes"].add(row.fe_id or row.candidate_fe)
        if row.service:
            item["services"].add(row.service)
        item["sources"].add(row.source_name)
        item["relations"].update(row.relation_tags)
        item["issues"].update(row.issue_tags)
    return [
        {
            "rr_id": rr_id,
            "fe_ids": ", ".join(sorted(item["fes"])),
            "service": ", ".join(sorted(item["services"])),
            "source": ", ".join(sorted(item["sources"])),
            "relation": ", ".join(sorted(item["relations"])) or "\u6b63\u5e38",
            "issue": ", ".join(sorted(item["issues"])) or "\u65e0",
        }
        for rr_id, item in sorted(grouped.items())
    ]


def _build_issue_rows(rows: list[NormalizedSourceRow]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for row in rows:
        for issue in sorted(row.issue_tags):
            output.append({
                "\u7ea7\u522b": row.severity,
                "\u5f02\u5e38": issue,
                "FE\u7f16\u53f7": row.fe_id,
                "RR\u7f16\u53f7": row.rr_id,
                "\u4e91\u670d\u52a1": row.service,
                "\u6765\u6e90\u5217\u8868": row.source_name,
                "\u884c\u53f7": row.row_number,
                "\u5904\u7406\u5efa\u8bae": _issue_suggestion(issue),
            })
    return output


def _issue_suggestion(issue: str) -> str:
    suggestions = {
        "\u7f3a\u4e3b\u952e": "\u8865\u5145 FE \u6216 RR \u540e\u91cd\u65b0\u5206\u6790",
        "\u7f3a\u4e91\u670d\u52a1\u5217": "\u4fee\u6b63\u6765\u6e90\u5217\u8868\u6216\u663e\u5f0f\u914d\u7f6e\u865a\u62df\u5217\u9ed8\u8ba4\u503c",
        "\u4e91\u670d\u52a1\u4e3a\u7a7a": "\u8865\u5145\u4e91\u670d\u52a1\u6216\u914d\u7f6e\u7a7a\u503c\u586b\u5145\u89c4\u5219",
        "\u5f85\u5efaFE": "\u5728\u5185\u90e8\u7cfb\u7edf\u521b\u5efa FE \u540e\u56de\u586b RR-FE \u6620\u5c04",
        "\u591a\u6765\u6e90\u5171\u540c\u7ba1\u7406": "\u786e\u8ba4\u662f\u5426\u91cd\u590d\u7eb3\u5165\u591a\u4e2a\u6765\u6e90\u5217\u8868",
        "\u4e91\u670d\u52a1\u51b2\u7a81": "\u786e\u8ba4 FE \u7684\u4e91\u670d\u52a1\u5f52\u5c5e",
    }
    return suggestions.get(issue, "\u5728\u8be6\u60c5\u4e2d\u786e\u8ba4\u540e\u5904\u7406")


def _build_extension_rows(rows: list[NormalizedSourceRow]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for row in rows:
        for field in row.extension_columns:
            value = normalize_cell(row.normalized.get(field))
            if value:
                output.append({
                    "\u6765\u6e90\u5217\u8868": row.source_name,
                    "FE\u7f16\u53f7": row.fe_id,
                    "RR\u7f16\u53f7": row.rr_id,
                    "\u4e91\u670d\u52a1": row.service,
                    "\u5b57\u6bb5": field,
                    "\u503c": value,
                })
    return output


def _build_pending_fe_rows(rows: list[NormalizedSourceRow]) -> list[dict[str, Any]]:
    return [
        {
            "RR\u7f16\u53f7": row.rr_id,
            "\u4e91\u670d\u52a1": row.service,
            "\u9700\u6c42\u6807\u9898": row.title,
            "\u6765\u6e90\u5217\u8868": row.source_name,
            "\u6765\u6e90\u7c7b\u578b": row.source_type,
            "\u521b\u5efa\u540eFE\u7f16\u53f7": "",
            "\u5904\u7406\u72b6\u6001": "\u5f85\u5efaFE",
            "\u5907\u6ce8": "",
        }
        for row in rows
        if "\u5f85\u5efaFE" in row.relation_tags
    ]


def _build_sync_comments(rows: list[NormalizedSourceRow], config: V2WorkflowConfig) -> list[dict[str, Any]]:
    comments: list[dict[str, Any]] = []
    for row in rows:
        if row.severity in {"\u4e25\u91cd", "\u9ad8\u4f18\u5148\u7ea7\u5f02\u5e38"}:
            continue
        selected_fields = list(config.comment_fields) or row.extension_columns
        if not any(normalize_cell(row.normalized.get(field)) for field in selected_fields):
            continue
        context = CommentContext(
            source_name=row.source_name,
            service=row.service,
            fe_id=row.fe_id or row.candidate_fe,
            rr_id=row.rr_id,
            title=row.title,
            row=row.normalized,
            selected_fields=selected_fields,
            extension_fields=row.extension_columns,
        )
        comment = render_comment(context, config.comment_template)
        for target_type, target_id in _comment_targets(row, config):
            comments.append({
                "\u76ee\u6807\u7c7b\u578b": target_type,
                "\u76ee\u6807ID": target_id,
                "FE\u7f16\u53f7": row.fe_id or row.candidate_fe,
                "RR\u7f16\u53f7": row.rr_id,
                "\u4e91\u670d\u52a1": row.service,
                "\u6765\u6e90\u5217\u8868": row.source_name,
                "\u8bc4\u8bba\u5185\u5bb9": comment,
                "\u72b6\u6001": "\u5f85\u786e\u8ba4",
            })
    return comments


def _comment_targets(row: NormalizedSourceRow, config: V2WorkflowConfig) -> list[tuple[str, str]]:
    fe_id = row.fe_id or row.candidate_fe
    rr_id = row.rr_id
    targets: list[tuple[str, str]] = []
    if config.comment_target in {"fe", "both"} and fe_id:
        targets.append(("FE", fe_id))
    if config.comment_target in {"rr", "both"} and rr_id:
        targets.append(("RR", rr_id))
    if not config.sync_linked:
        return _unique_targets(targets)
    if fe_id and rr_id:
        targets.extend([("FE", fe_id), ("RR", rr_id)])
    return _unique_targets(targets)


def _unique_targets(targets: list[tuple[str, str]]) -> list[tuple[str, str]]:
    seen: set[tuple[str, str]] = set()
    output: list[tuple[str, str]] = []
    for target in targets:
        if target not in seen:
            seen.add(target)
            output.append(target)
    return output


def _build_summary(preview_rows: list[dict[str, Any]], rr_rows: list[dict[str, Any]], rows: list[NormalizedSourceRow], sync_comments: list[dict[str, Any]]) -> dict[str, Any]:
    services = {row["service"] for row in preview_rows if row["service"]}
    sources = {row.source_name for row in rows}
    return {
        "fe_count": len({fe for row in preview_rows for fe in row["fe_id"].split(", ") if fe}),
        "rr_count": len({rr for row in rr_rows for rr in [row["rr_id"]] if rr}),
        "service_count": len(services),
        "source_list_count": len(sources),
        "pending_fe_count": sum(1 for row in rows if "\u5f85\u5efaFE" in row.relation_tags),
        "candidate_fe_count": sum(1 for row in rows if "FE\u5019\u9009" in row.relation_tags),
        "multi_source_count": sum(1 for row in preview_rows if "\u591a\u6765\u6e90\u5171\u540c\u7ba1\u7406" in row["relation"]),
        "critical_issue_count": sum(1 for row in rows if row.severity == "\u4e25\u91cd"),
        "sync_comment_count": len(sync_comments),
        "service_distribution": _service_distribution(preview_rows),
    }


def _service_distribution(preview_rows: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in preview_rows:
        for service in [item.strip() for item in row["service"].split(",") if item.strip()]:
            counts[service] = counts.get(service, 0) + 1
    return counts


def _build_filters(preview_rows: list[dict[str, Any]]) -> dict[str, list[str]]:
    return {
        "services": sorted({row["service"] for row in preview_rows if row["service"]}),
        "sources": sorted({source.strip() for row in preview_rows for source in row["source"].split(",") if source.strip()}),
        "severities": sorted({row["severity"] for row in preview_rows if row["severity"]}),
        "owned": ["\u5168\u90e8", "\u662f", "\u5426"],
    }


def _build_raw_rows(rows: list[NormalizedSourceRow], config: V2WorkflowConfig) -> list[dict[str, Any]]:
    all_fields: list[str] = []
    for row in rows:
        for field in row.raw:
            if field not in all_fields:
                all_fields.append(field)
    output: list[dict[str, Any]] = []
    for row in rows:
        item = {
            "\u6765\u6e90\u5217\u8868": row.source_name,
            "\u884c\u53f7": row.row_number,
            "\u586b\u5145\u8bf4\u660e": "; ".join(row.fill_notes),
        }
        for field in all_fields:
            item[field] = row.raw.get(field)
        output.append(item)
    return output


def _write_package(
    package_path: Path,
    summary: dict[str, Any],
    preview_rows: list[dict[str, Any]],
    rr_rows: list[dict[str, Any]],
    extension_rows: list[dict[str, Any]],
    sync_comments: list[dict[str, Any]],
    pending_fe_rows: list[dict[str, Any]],
    issue_rows: list[dict[str, Any]],
    raw_rows: list[dict[str, Any]],
) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "\u603b\u89c8"
    _write_key_value_sheet(summary_sheet, summary)
    _write_rows(workbook.create_sheet("FE\u89c6\u89d2\u9700\u6c42\u6e05\u5355"), preview_rows)
    _write_rows(workbook.create_sheet("RR\u89c6\u89d2\u9700\u6c42\u6e05\u5355"), rr_rows)
    _write_rows(workbook.create_sheet("\u4e13\u9879\u89c6\u89d2\u9700\u6c42\u6e05\u5355"), preview_rows)
    _write_rows(workbook.create_sheet("\u6211\u8d1f\u8d23\u7684\u4e91\u670d\u52a1\u9700\u6c42"), [row for row in preview_rows if row["owned"] == "\u662f"])
    _write_rows(workbook.create_sheet("\u6765\u6e90\u5217\u8868\u6269\u5c55\u5b57\u6bb5"), extension_rows)
    _write_rows(workbook.create_sheet("\u5f85\u56de\u5237\u8bc4\u8bba"), sync_comments)
    _write_rows(workbook.create_sheet("\u5f85\u5efaFE"), pending_fe_rows)
    _write_rows(workbook.create_sheet("\u5f02\u5e38\u9879"), issue_rows)
    _write_rows(workbook.create_sheet("\u539f\u59cb\u884c\u6570\u636e"), raw_rows)
    workbook.save(package_path)


def _write_key_value_sheet(sheet: Worksheet, summary: dict[str, Any]) -> None:
    _write_header(sheet, ["\u6307\u6807", "\u503c"])
    for key, value in summary.items():
        if isinstance(value, dict):
            sheet.append([key, json.dumps(value, ensure_ascii=False)])
        else:
            sheet.append([key, value])
    _autosize_columns(sheet)


def _write_rows(sheet: Worksheet, rows: list[dict[str, Any]]) -> None:
    if not rows:
        _write_header(sheet, ["\u6682\u65e0\u6570\u636e"])
        _autosize_columns(sheet)
        return
    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    _write_header(sheet, headers)
    for row in rows:
        sheet.append([_excel_value(row.get(header)) for header in headers])
    _autosize_columns(sheet)


def _excel_value(value: Any) -> Any:
    if isinstance(value, (dict, list, set, tuple)):
        return json.dumps(list(value) if isinstance(value, set) else value, ensure_ascii=False)
    return value


def _write_header(sheet: Worksheet, headers: list[str]) -> None:
    sheet.append(headers)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill


def _autosize_columns(sheet: Worksheet) -> None:
    for column_cells in sheet.columns:
        max_length = 10
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 80))
        sheet.column_dimensions[column_letter].width = max_length + 2

