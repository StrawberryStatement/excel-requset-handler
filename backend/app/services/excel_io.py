from __future__ import annotations

import json
from dataclasses import asdict, is_dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from backend.app.models import CellValue, ExcelTable, PmAnalysis, SyncComment


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_cell(value: CellValue) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_excel_table(path: Path, sheet_name: str | None = None, header_row: int = 1) -> ExcelTable:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    headers = [normalize_header(cell.value) for cell in sheet[header_row]]
    if not any(headers):
        raise ValueError(f"{path.name} does not contain headers in row {header_row}")
    _validate_headers(headers, path.name)

    records: list[dict[str, CellValue]] = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        record: dict[str, CellValue] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            record[header] = row[index] if index < len(row) else None
        records.append(record)

    return ExcelTable(
        source_name=path.stem,
        sheet_name=sheet.title,
        headers=[header for header in headers if header],
        records=records,
    )


def _validate_headers(headers: list[str], source_name: str) -> None:
    seen: set[str] = set()
    duplicates: list[str] = []
    for header in headers:
        if not header:
            continue
        if header in seen and header not in duplicates:
            duplicates.append(header)
        seen.add(header)

    if duplicates:
        joined = ", ".join(duplicates)
        raise ValueError(f"Duplicate header in {source_name}: {joined}")


def write_table(path: Path, sheet_name: str, headers: list[str], rows: Iterable[dict[str, CellValue]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    _write_header(sheet, headers)

    for row in rows:
        sheet.append([row.get(header) for header in headers])

    _autosize_columns(sheet)
    workbook.save(path)


def write_diff_report(path: Path, analyses: list[PmAnalysis]) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "summary"
    _write_header(summary, ["source_file", "extension_column_count", "baseline_change_count", "issue_count"])
    for analysis in analyses:
        summary.append([
            analysis.source_name,
            len(analysis.extension_columns),
            len(analysis.changed_baseline_fields),
            len(analysis.issues),
        ])

    changed_sheet = workbook.create_sheet("changed_baseline_fields")
    _write_header(changed_sheet, ["source_file", "fe_id", "field", "baseline_value", "pm_value"])
    for analysis in analyses:
        for change in analysis.changed_baseline_fields:
            changed_sheet.append([
                change.source_name,
                change.fe_id,
                change.field,
                change.baseline_value,
                change.pm_value,
            ])

    extension_sheet = workbook.create_sheet("extension_columns")
    _write_header(extension_sheet, ["source_file", "extension_column"])
    for analysis in analyses:
        for column in analysis.extension_columns:
            extension_sheet.append([analysis.source_name, column])

    issues_sheet = workbook.create_sheet("issues")
    _write_header(issues_sheet, ["level", "source_file", "fe_id", "message", "row_number"])
    for analysis in analyses:
        for issue in analysis.issues:
            issues_sheet.append([
                issue.level,
                issue.source_name,
                issue.fe_id,
                issue.message,
                issue.row_number,
            ])

    for sheet in workbook.worksheets:
        _autosize_columns(sheet)
    workbook.save(path)


def write_sync_plan_json(path: Path, comments: list[SyncComment]) -> None:
    payload = [_to_plain_dict(comment) for comment in comments]
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_sync_plan_workbook(path: Path, comments: list[SyncComment]) -> None:
    headers = ["FE\u7f16\u53f7", "\u9700\u6c42\u6807\u9898", "\u8bc4\u8bba\u5185\u5bb9", "\u6765\u6e90\u6587\u4ef6", "\u72b6\u6001"]
    rows = [
        {
            "FE\u7f16\u53f7": comment.fe_id,
            "\u9700\u6c42\u6807\u9898": comment.title,
            "\u8bc4\u8bba\u5185\u5bb9": comment.comment,
            "\u6765\u6e90\u6587\u4ef6": ", ".join(comment.source_names),
            "\u72b6\u6001": comment.status,
        }
        for comment in comments
    ]
    write_table(path, "sync_plan", headers, rows)


def _write_header(sheet: Worksheet, headers: list[str]) -> None:
    sheet.append(headers)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill


def _autosize_columns(sheet: Worksheet) -> None:
    for column_cells in sheet.columns:
        max_length = 10
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 60))
        sheet.column_dimensions[column_letter].width = max_length + 2


def _to_plain_dict(value: object) -> object:
    if is_dataclass(value):
        return {key: _to_plain_dict(item) for key, item in asdict(value).items()}
    if isinstance(value, list):
        return [_to_plain_dict(item) for item in value]
    return value
