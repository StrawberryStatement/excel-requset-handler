from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.app.models import WorkbookConfig
from backend.app.services.workflow import run_excel_workflow
from tests.conftest import CONCLUSION, FE, RISK, SOURCE_PM, STATUS, write_workbook


def test_workflow_generates_all_outputs(tmp_path: Path, baseline_headers, baseline_rows) -> None:
    baseline = tmp_path / "baseline.xlsx"
    pm = tmp_path / "pm_alpha.xlsx"
    write_workbook(baseline, baseline_headers, baseline_rows)
    write_workbook(
        pm,
        baseline_headers + [RISK, CONCLUSION],
        [["FE-1", "RR-1", "A", "Login polish", "Doing", "2026Q2", "Dev A", "2026-05-09", "Gateway", "Approved"]],
    )

    result = run_excel_workflow(baseline, [pm], tmp_path / "jobs")

    assert result.master_workbook_path.exists()
    assert result.diff_report_path.exists()
    assert result.sync_plan_json_path.exists()
    assert result.sync_plan_workbook_path.exists()
    assert (result.output_dir / "job_summary.json").exists()

    master_workbook = load_workbook(result.master_workbook_path)
    master_headers = [cell.value for cell in master_workbook.active[1]]
    assert "pm_alpha__" + RISK in master_headers

    diff_workbook = load_workbook(result.diff_report_path)
    assert "changed_baseline_fields" in diff_workbook.sheetnames

    sync_payload = json.loads(result.sync_plan_json_path.read_text(encoding="utf-8"))
    assert sync_payload[0]["fe_id"] == "FE-1"
    assert RISK in sync_payload[0]["comment"]

    summary = json.loads((result.output_dir / "job_summary.json").read_text(encoding="utf-8"))
    assert summary["job_id"] == result.job_id
    assert summary["baseline_change_count"] == 1
    assert summary["sync_comment_count"] == 1
    assert summary["issue_count"] == 0


def test_workflow_rejects_missing_fe_column(tmp_path: Path, baseline_headers, baseline_rows) -> None:
    baseline = tmp_path / "baseline.xlsx"
    pm = tmp_path / "pm_alpha.xlsx"
    write_workbook(baseline, [header for header in baseline_headers if header != FE], [row[1:] for row in baseline_rows])
    write_workbook(pm, baseline_headers, baseline_rows)

    with pytest.raises(ValueError, match="missing required columns"):
        run_excel_workflow(baseline, [pm], tmp_path / "jobs")


def test_workflow_rejects_duplicate_baseline_fe(tmp_path: Path, baseline_headers, baseline_rows) -> None:
    baseline = tmp_path / "baseline.xlsx"
    pm = tmp_path / "pm_alpha.xlsx"
    duplicated_rows = baseline_rows + [["FE-1", "RR-X", "A", "Duplicate", "Review", "2026Q4", "Dev X", "2026-05-09"]]
    write_workbook(baseline, baseline_headers, duplicated_rows)
    write_workbook(pm, baseline_headers, baseline_rows)

    with pytest.raises(ValueError, match="Duplicate baseline FE"):
        run_excel_workflow(baseline, [pm], tmp_path / "jobs")


def test_workflow_supports_custom_key_column_names(tmp_path: Path) -> None:
    baseline = tmp_path / "baseline.xlsx"
    pm = tmp_path / "pm_alpha.xlsx"
    headers = ["FeatureID", "RR", "Title", STATUS]
    write_workbook(baseline, headers, [["FE-1", "RR-1", "Login polish", "Review"]])
    write_workbook(pm, headers + [CONCLUSION], [["FE-1", "RR-1", "Login polish", "Doing", "Approved"]])

    result = run_excel_workflow(
        baseline,
        [pm],
        tmp_path / "jobs",
        WorkbookConfig(fe_column="FeatureID", rr_column="RR", title_column="Title"),
    )

    assert result.analyses[0].changed_baseline_fields[0].field == STATUS
    assert result.sync_comments[0].fe_id == "FE-1"


def test_workflow_aggregates_multiple_pm_files_for_same_fe(tmp_path: Path, baseline_headers, baseline_rows) -> None:
    baseline = tmp_path / "baseline.xlsx"
    pm_a = tmp_path / "pm_a.xlsx"
    pm_b = tmp_path / "pm_b.xlsx"
    write_workbook(baseline, baseline_headers, baseline_rows)
    write_workbook(pm_a, baseline_headers + [RISK], [["FE-1", "RR-1", "A", "Login polish", "Review", "2026Q2", "Dev A", "2026-05-01", "Gateway"]])
    write_workbook(pm_b, baseline_headers + [CONCLUSION], [["FE-1", "RR-1", "A", "Login polish", "Review", "2026Q2", "Dev A", "2026-05-01", "Aligned"]])

    result = run_excel_workflow(baseline, [pm_a, pm_b], tmp_path / "jobs")

    workbook = load_workbook(result.master_workbook_path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    row = {headers[index]: cell.value for index, cell in enumerate(sheet[2])}

    assert row[SOURCE_PM] == "pm_a, pm_b"
    assert row["pm_a__" + RISK] == "Gateway"
    assert row["pm_b__" + CONCLUSION] == "Aligned"
    assert len(result.sync_comments) == 1
    assert "[pm_a]" in result.sync_comments[0].comment
    assert "[pm_b]" in result.sync_comments[0].comment
