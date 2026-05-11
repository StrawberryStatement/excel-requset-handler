from __future__ import annotations

import json

from openpyxl import load_workbook

from backend.app.services.v2_workflow import V2WorkflowConfig, run_v2_workflow
from tests.conftest import CONCLUSION, FE, RR, RISK, SERVICE, STATUS, TITLE, write_workbook


def test_v2_workflow_generates_preview_package_and_comments(tmp_path) -> None:
    baseline = tmp_path / "baseline.xlsx"
    source = tmp_path / "专项1需求列表.xlsx"
    output = tmp_path / "jobs"
    headers = [FE, RR, SERVICE, TITLE, STATUS]
    write_workbook(
        baseline,
        headers,
        [
            ["FE-1", "RR-1", "A", "Login polish", "Doing"],
            ["FE-2", "RR-2", "B", "Export report", "Planning"],
        ],
    )
    write_workbook(
        source,
        headers + [CONCLUSION, RISK],
        [["FE-1", "RR-1", "A", "Login polish", "Doing", "与研发对齐按Q2交付", "低风险"]],
    )

    result = run_v2_workflow(
        baseline_path=baseline,
        source_paths=[source],
        output_root=output,
        config=V2WorkflowConfig(
            owned_services=("A",),
            comment_fields=(CONCLUSION, RISK),
            comment_target="both",
        ),
    )

    payload = result.response_payload
    assert payload["summary"]["fe_count"] == 2
    assert payload["summary"]["source_list_count"] == 1
    assert payload["summary"]["sync_comment_count"] == 2
    assert any(row["fe_id"] == "FE-1" and row["owned"] == "是" for row in payload["preview_rows"])
    assert any(item["目标类型"] == "FE" and item["目标ID"] == "FE-1" for item in _read_sync_plan(payload, result))
    assert any(item["目标类型"] == "RR" and item["目标ID"] == "RR-1" for item in _read_sync_plan(payload, result))

    workbook = load_workbook(result.package_path)
    assert {
        "总览",
        "FE视角需求清单",
        "RR视角需求清单",
        "待回刷评论",
        "待建FE",
        "异常项",
    }.issubset(set(workbook.sheetnames))


def test_v2_workflow_marks_pending_fe_and_multi_source(tmp_path) -> None:
    baseline = tmp_path / "baseline.xlsx"
    source_a = tmp_path / "专项1需求列表.xlsx"
    source_b = tmp_path / "服务A需求列表.xlsx"
    output = tmp_path / "jobs"
    headers = [FE, RR, SERVICE, TITLE]
    write_workbook(
        baseline,
        headers,
        [
            ["FE-1", "RR-1", "A", "Login polish"],
            ["FE-2", "RR-2", "A", "Export report"],
        ],
    )
    write_workbook(
        source_a,
        headers + [CONCLUSION],
        [
            ["FE-1", "RR-1", "A", "Login polish", "专项侧需要高亮"],
            ["", "RR-NEW", "C", "New customer ask", "需要新建FE"],
        ],
    )
    write_workbook(
        source_b,
        headers + [CONCLUSION],
        [["FE-1", "RR-1", "A", "Login polish", "服务侧也在管理"]],
    )

    result = run_v2_workflow(
        baseline_path=baseline,
        source_paths=[source_a, source_b],
        output_root=output,
        config=V2WorkflowConfig(comment_fields=(CONCLUSION,)),
    )

    payload = result.response_payload
    assert payload["summary"]["pending_fe_count"] == 1
    assert payload["summary"]["multi_source_count"] == 1
    assert any("待建FE" in row["relation"] for row in payload["rr_rows"])
    assert any("多来源共同管理" in row["issue"] for row in payload["preview_rows"])


def test_v2_workflow_marks_duplicate_baseline_fe_without_rejecting(tmp_path) -> None:
    baseline = tmp_path / "baseline.xlsx"
    source = tmp_path / "专项1需求列表.xlsx"
    output = tmp_path / "jobs"
    headers = [FE, RR, SERVICE, TITLE]
    write_workbook(
        baseline,
        headers,
        [
            ["FE-1", "RR-1", "A", "Login polish"],
            ["FE-1", "RR-2", "A", "Login polish duplicate"],
        ],
    )
    write_workbook(
        source,
        headers + [CONCLUSION],
        [["FE-1", "RR-1", "A", "Login polish", "需要确认基准重复"]],
    )

    result = run_v2_workflow(
        baseline_path=baseline,
        source_paths=[source],
        output_root=output,
        config=V2WorkflowConfig(comment_fields=(CONCLUSION,)),
    )

    payload = result.response_payload
    assert payload["summary"]["critical_issue_count"] >= 1
    assert any("基准重复FE" in row["issue"] for row in payload["preview_rows"])
    assert any(item["异常"] == "基准重复FE" for item in payload["issues"])


def test_v2_workflow_matches_owned_services_case_insensitively(tmp_path) -> None:
    baseline = tmp_path / "baseline.xlsx"
    source = tmp_path / "专项1需求列表.xlsx"
    output = tmp_path / "jobs"
    headers = [FE, RR, SERVICE, TITLE]
    write_workbook(baseline, headers, [["FE-1", "RR-1", "clouda", "Login polish"]])
    write_workbook(source, headers + [CONCLUSION], [["FE-1", "RR-1", "CloudA", "Login polish", "已对齐"]])

    result = run_v2_workflow(
        baseline_path=baseline,
        source_paths=[source],
        output_root=output,
        config=V2WorkflowConfig(owned_services=("clouda",), comment_fields=(CONCLUSION,)),
    )

    assert any(row["fe_id"] == "FE-1" and row["owned"] == "是" for row in result.response_payload["preview_rows"])


def _read_sync_plan(payload: dict, result) -> list[dict]:
    assert payload["downloads"]["sync_plan_json"].endswith("/sync_plan.json")
    return json.loads(result.sync_plan_json_path.read_text(encoding="utf-8"))
